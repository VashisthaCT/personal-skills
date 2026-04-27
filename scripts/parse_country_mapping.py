#!/usr/bin/env python3
"""
parse_country_mapping.py — extract field mapping + enums + open queries from a
consultant's xlsx (Avtax-style spreadsheets) into structured YAML.

Used by `/v-country-onboard parse-mapping <cc> --xlsx <path>`.

Outputs (to ~/dev/personal-skills/data/onboarding/<cc>/):
  - field_mapping.yaml — api_path, db_path, xml_xpath, mandatory, conditions per row
  - enums.yaml         — sub-type codes, buyer ID schemes, tax categories, governorates
  - open_queries.md    — spec conflicts and to-be-tested rows

Sheet detection is fuzzy because consultants name them differently per country:
  - "(Tech)JORDAN field to api and d" → field mapping
  - "Field to API mapping" → field mapping
  - "Enums" → enums
  - "Open Queries" → open queries

CLI:
    python3 scripts/parse_country_mapping.py --country EG --xlsx ~/Downloads/Egypt_xx.xlsx

Stdlib + openpyxl. If openpyxl missing, prints install command and exits cleanly.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Any


REPO_ROOT = Path(__file__).resolve().parent.parent


# ---------- openpyxl import with friendly fallback ----------

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed.", file=sys.stderr)
    print("Install: pip3 install --user openpyxl", file=sys.stderr)
    sys.exit(2)


# ---------- sheet name fuzzy matchers ----------

FIELD_MAPPING_KEYWORDS = ("tech", "field", "api", "mapping", "ubl")
ENUMS_KEYWORDS = ("enum",)
OPEN_QUERIES_KEYWORDS = ("open", "queries", "questions", "todo")


def _score_sheet_name(name: str, keywords: tuple[str, ...]) -> int:
    """Count keyword hits in a sheet name (case-insensitive)."""
    lower = name.lower()
    return sum(1 for kw in keywords if kw in lower)


def find_tech_sheet(workbook) -> Any | None:
    """Fuzzy-find the field-to-api mapping sheet."""
    candidates = []
    for name in workbook.sheetnames:
        score = _score_sheet_name(name, FIELD_MAPPING_KEYWORDS)
        if score >= 2:
            candidates.append((score, name))
    if not candidates:
        return None
    candidates.sort(reverse=True)
    return workbook[candidates[0][1]]


def find_enums_sheet(workbook) -> Any | None:
    """Fuzzy-find the enums sheet."""
    for name in workbook.sheetnames:
        if _score_sheet_name(name, ENUMS_KEYWORDS) >= 1:
            return workbook[name]
    return None


def find_open_queries_sheet(workbook) -> Any | None:
    """Fuzzy-find the open-queries sheet."""
    candidates = []
    for name in workbook.sheetnames:
        score = _score_sheet_name(name, OPEN_QUERIES_KEYWORDS)
        if score >= 1:
            candidates.append((score, name))
    if not candidates:
        return None
    candidates.sort(reverse=True)
    return workbook[candidates[0][1]]


# ---------- header detection ----------

API_PATH_HEADERS = ("api", "field", "json path", "json", "path", "pint")
DB_PATH_HEADERS = ("db", "ubl", "internal", "model")
XML_HEADERS = ("xpath", "xml", "ubl xpath")
MANDATORY_HEADERS = ("mandatory", "required", "m/o", "m / c")
CONDITION_HEADERS = ("condition", "rule", "note", "remark")


def _normalize_header(s: Any) -> str:
    if s is None:
        return ""
    return str(s).strip().lower()


def _classify_column(header: str) -> str | None:
    """Map a header cell to a normalized column key, or None if unrecognized."""
    h = _normalize_header(header)
    if not h:
        return None
    for kw in API_PATH_HEADERS:
        if kw in h:
            return "api_path"
    for kw in DB_PATH_HEADERS:
        if kw in h:
            return "db_path"
    for kw in XML_HEADERS:
        if kw in h:
            return "xml_xpath"
    for kw in MANDATORY_HEADERS:
        if kw in h:
            return "mandatory"
    for kw in CONDITION_HEADERS:
        if kw in h:
            return "condition"
    return None


def _find_header_row(sheet, max_rows: int = 10) -> int | None:
    """Find the first row that has at least 2 recognized header cells."""
    for row_idx in range(1, max_rows + 1):
        header_hits = 0
        for cell in sheet[row_idx]:
            if _classify_column(cell.value) is not None:
                header_hits += 1
        if header_hits >= 2:
            return row_idx
    return None


# ---------- parsing ----------

def parse_field_mapping(sheet) -> list[dict[str, Any]]:
    """Walk the sheet's rows after the header row; emit one dict per data row."""
    if sheet is None:
        return []

    header_row = _find_header_row(sheet)
    if header_row is None:
        print(f"WARN: could not find header row in '{sheet.title}'", file=sys.stderr)
        return []

    column_map: dict[int, str] = {}
    for cell in sheet[header_row]:
        key = _classify_column(cell.value)
        if key is not None:
            column_map[cell.column] = key

    rows = []
    for row in sheet.iter_rows(min_row=header_row + 1, values_only=False):
        rec: dict[str, Any] = {}
        for cell in row:
            key = column_map.get(cell.column)
            if key is None:
                continue
            value = cell.value
            if isinstance(value, str):
                value = value.strip()
            if value is not None and value != "":
                rec[key] = value
        if rec.get("api_path") or rec.get("db_path") or rec.get("xml_xpath"):
            rows.append(rec)
    return rows


def parse_enums(sheet) -> dict[str, list[Any]]:
    """Best-effort: each sub-table on the enums sheet becomes a list under a key."""
    if sheet is None:
        return {}

    out: dict[str, list[Any]] = {}
    current_key: str | None = None
    current_rows: list[Any] = []

    for row in sheet.iter_rows(values_only=True):
        # Heuristic: a single non-empty cell on a row = a section title.
        non_empty = [v for v in row if v is not None and str(v).strip() != ""]
        if len(non_empty) == 1:
            if current_key and current_rows:
                out[current_key] = current_rows
            current_key = str(non_empty[0]).strip().lower().replace(" ", "_")
            current_rows = []
            continue
        if len(non_empty) >= 2 and current_key is not None:
            current_rows.append([str(v).strip() if v is not None else "" for v in row if v is not None])

    if current_key and current_rows:
        out[current_key] = current_rows
    return out


def parse_open_queries(sheet) -> list[dict[str, Any]]:
    """Each non-empty row is a query; first cell = ID, second = description."""
    if sheet is None:
        return []
    queries = []
    for row in sheet.iter_rows(values_only=True):
        non_empty = [v for v in row if v is not None and str(v).strip() != ""]
        if not non_empty:
            continue
        if len(non_empty) == 1:
            queries.append({"id": "", "text": str(non_empty[0]).strip()})
        else:
            queries.append({
                "id": str(non_empty[0]).strip(),
                "text": " | ".join(str(v).strip() for v in non_empty[1:]),
            })
    return queries


# ---------- YAML serialization (no external dep) ----------

def _yaml_value(v: Any, indent: int = 0) -> str:
    """Serialize a Python value to YAML — minimal, stdlib-only, no PyYAML dependency."""
    pad = "  " * indent
    if v is None:
        return "null"
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (int, float)):
        return str(v)
    if isinstance(v, str):
        # Quote if contains special chars or starts with special punctuation
        needs_quote = any(c in v for c in ":#\n'\"") or v.strip() != v or v == ""
        if needs_quote:
            escaped = v.replace("\\", "\\\\").replace('"', '\\"').replace("\n", "\\n")
            return f'"{escaped}"'
        return v
    if isinstance(v, list):
        if not v:
            return "[]"
        lines = []
        for item in v:
            if isinstance(item, dict):
                lines.append(f"{pad}-")
                for k, sub in item.items():
                    lines.append(f"{pad}  {k}: {_yaml_value(sub, indent + 2)}")
            elif isinstance(item, list):
                lines.append(f"{pad}- {_yaml_value(item, indent + 1)}")
            else:
                lines.append(f"{pad}- {_yaml_value(item, indent + 1)}")
        return "\n" + "\n".join(lines)
    if isinstance(v, dict):
        if not v:
            return "{}"
        lines = []
        for k, sub in v.items():
            lines.append(f"{pad}{k}: {_yaml_value(sub, indent + 1)}")
        return "\n" + "\n".join(lines)
    return str(v)


def write_yaml(parsed: dict[str, Any], output_path: Path) -> None:
    """Write parsed dict as YAML to output_path."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w") as f:
        for k, v in parsed.items():
            f.write(f"{k}: {_yaml_value(v, 1)}\n")


def write_open_queries_md(queries: list[dict[str, Any]], output_path: Path) -> None:
    """Write open queries as a markdown checklist."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "# Open Queries",
        "",
        "Auto-extracted from the consultant's xlsx Open Queries sheet.",
        "Edit before sending to consultant. Skill `/v-country-onboard init` will ingest answers.",
        "",
    ]
    for q in queries:
        prefix = f"**{q['id']}**: " if q.get("id") else ""
        lines.append(f"- [ ] {prefix}{q.get('text', '')}")
    output_path.write_text("\n".join(lines) + "\n")


# ---------- main ----------

def main() -> int:
    parser = argparse.ArgumentParser(
        description="Extract field mapping + enums + open queries from a consultant's xlsx."
    )
    parser.add_argument("--country", required=True, help="2-letter country code (e.g. EG, OM)")
    parser.add_argument("--xlsx", required=True, help="Path to the consultant's xlsx")
    parser.add_argument(
        "--out-dir",
        default=None,
        help="Output dir (default: ~/dev/personal-skills/data/onboarding/<cc>/)",
    )
    parser.add_argument(
        "--sheet-name",
        default=None,
        help="Override fuzzy detection — exact field-mapping sheet name",
    )
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx).expanduser()
    if not xlsx_path.exists():
        print(f"ERROR: xlsx not found at {xlsx_path}", file=sys.stderr)
        return 1

    cc_lower = args.country.lower()
    if args.out_dir:
        out_dir = Path(args.out_dir).expanduser()
    else:
        out_dir = REPO_ROOT / "data" / "onboarding" / cc_lower

    print(f"Loading {xlsx_path} …", file=sys.stderr)
    workbook = load_workbook(xlsx_path, data_only=True, read_only=True)

    # Field mapping
    if args.sheet_name:
        if args.sheet_name not in workbook.sheetnames:
            print(f"ERROR: sheet '{args.sheet_name}' not found.", file=sys.stderr)
            print(f"Available: {workbook.sheetnames}", file=sys.stderr)
            return 1
        tech_sheet = workbook[args.sheet_name]
    else:
        tech_sheet = find_tech_sheet(workbook)

    if tech_sheet is None:
        print("WARN: no tech/field-mapping sheet detected.", file=sys.stderr)
        print(f"Sheets in workbook: {workbook.sheetnames}", file=sys.stderr)
        print("Re-run with --sheet-name <exact-name> to override.", file=sys.stderr)
    else:
        print(f"Tech sheet: {tech_sheet.title}", file=sys.stderr)

    field_mapping = parse_field_mapping(tech_sheet) if tech_sheet else []

    # Enums
    enums_sheet = find_enums_sheet(workbook)
    if enums_sheet:
        print(f"Enums sheet: {enums_sheet.title}", file=sys.stderr)
    enums = parse_enums(enums_sheet) if enums_sheet else {}

    # Open queries
    queries_sheet = find_open_queries_sheet(workbook)
    if queries_sheet:
        print(f"Open queries sheet: {queries_sheet.title}", file=sys.stderr)
    queries = parse_open_queries(queries_sheet) if queries_sheet else []

    # Write outputs
    write_yaml({"country": args.country.upper(), "field_mapping": field_mapping}, out_dir / "field_mapping.yaml")
    write_yaml({"country": args.country.upper(), "enums": enums}, out_dir / "enums.yaml")
    write_open_queries_md(queries, out_dir / "open_queries.md")

    print(f"Wrote {len(field_mapping)} field rows → {out_dir / 'field_mapping.yaml'}")
    print(f"Wrote {len(enums)} enum sections → {out_dir / 'enums.yaml'}")
    print(f"Wrote {len(queries)} open queries → {out_dir / 'open_queries.md'}")
    print(f"Next: /v-country-onboard init {cc_lower}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
